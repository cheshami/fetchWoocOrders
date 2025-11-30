import json
import subprocess
import os
import logging
import requests
import jdatetime
import concurrent.futures
from woocommerce import API
from typing import List, Any
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from datetime import datetime
from logging.handlers import RotatingFileHandler

def get_key_by_value(d, value):
    """Return the key associated with the given value in the dictionary."""
    for key, val in d.items():
        if val == value:
            return key
    return None

def find_meta_value(data, target_key):
    """Function to find the value for a specific key in meta_data"""
    for item in data.get('meta_data', []):
        if item['key'] == target_key:
            return item['value']
    return None

def validate_config(config):
    """Validate the loaded configuration."""
    required_keys = {
        "WC_API": dict,
        "color": dict,
        "borders": dict,
        "font": dict
    }

    for key, expected_type in required_keys.items():
        if key not in config:
            logging.error(f"Missing required configuration key: {key}")
            raise ValueError(f"Missing required configuration key: {key}")

        if not isinstance(config[key], expected_type):
            logging.error(f"Incorrect type for key '{key}': expected {expected_type.__name__}, got {type(config[key]).__name__}")
            raise ValueError(f"Incorrect type for key '{key}': expected {expected_type.__name__}, got {type(config[key]).__name__}")

    logging.info("Configuration validated successfully.")

def load_config(json_file):
    """Load configuration from a JSON file."""
    try:
        with open(json_file, 'r') as file:
            config = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"Error loading config: {e}")
        raise

    # Validate the configuration
    validate_config(config)

    wc_api = config["WC_API"]
    days_to_fetch = config["DAYS"]

    color = {name: PatternFill(**config["color"][name]) for name in config["color"]}
    
    borders = {
        "all_borders": Border(**{side: Side(style=config["borders"]["thin_border"][side]) for side in ["left", "right", "top", "bottom"]}),
        "left_border": Border(left=Side(style=config["borders"]["thin_border"]["left"])),
        "right_border": Border(right=Side(style=config["borders"]["thin_border"]["right"])),
        "top_border": Border(top=Side(style=config["borders"]["thin_border"]["top"])),
        "bottom_border": Border(bottom=Side(style=config["borders"]["thin_border"]["bottom"]))
    }

    fonts = {name: Font(name=config["font"][name]) for name in config["font"]}

    return {"lang": config['lang']}, wc_api, days_to_fetch, color, borders, fonts

def convert_to_jalali(date_string):
    """Convert ISO date string to Jalali date format."""
    date_part, time_part = date_string.split('T')
    gregorian_date = datetime.fromisoformat(date_part)
    jalali_date = jdatetime.date.fromgregorian(year=gregorian_date.year,
                                               month=gregorian_date.month, day=gregorian_date.day)
    return jalali_date, time_part

def apply_styles(sheet, row_index, order):
    """Apply styles to a specific row in the sheet."""
    for cell in sheet[row_index]:
        cell.fill = color["order_bg"]
        cell.border = borders["all_borders"]

    if int(order['discount_total']) > 0:
        sheet[row_index][list(COLUMN_HEADERS.keys()).index("sepidar_discount")].fill = color["red_bg"]

    for col in ["total", "shipping", "discount", "sepidar_discount", "com_postal_payment", "com_postage"]:
        sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index(col) + 1).number_format = '#,0'

    for col in ["date_paid", "phone", "postcode"]:
        sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index(col) + 1).number_format = '@'

def fetch_page(master_page: str, params: dict, max_retries: int = 3) -> list:
    """Fetch orders from WooCommerce API from a specific page with retry logic and increasing timeouts."""
    for attempt in range(max_retries):
        try:
            # Increase timeout by 5 seconds for each retry attempt
            current_timeout = wc_api['TIMEOUT'] + (5 * attempt)
            
            wcapi = API(
                url=wc_api['URL'],
                consumer_key=wc_api['KEY'],
                consumer_secret=wc_api['SECRET'],
                wp_api=True,
                version="wc/v3",
                timeout=current_timeout,  # Use the increased timeout
            )
            response = wcapi.get(master_page, params=params)
            response.raise_for_status()
            return response.json()
        
        except requests.exceptions.Timeout:
            if attempt < max_retries - 1:
                current_timeout = wc_api['TIMEOUT'] + (5 * (attempt + 1))
                logging.warning(f"Timeout on page {params['page']}, retry {attempt + 1}/{max_retries} with increased timeout {current_timeout}s")
                continue
            else:
                logging.error(f"Timeout after {max_retries} attempts on page {params['page']}")
                return []
        
        except requests.exceptions.ConnectionError:
            if attempt < max_retries - 1:
                current_timeout = wc_api['TIMEOUT'] + (5 * (attempt + 1))
                logging.warning(f"Connection error on page {params['page']}, retry {attempt + 1}/{max_retries} with increased timeout {current_timeout}s")
                continue
            else:
                logging.error(f"Connection error after {max_retries} attempts on page {params['page']}")
                return []
        
        except requests.exceptions.HTTPError as http_err:
            logging.error(f"HTTP error occurred on page {params['page']}: {http_err}")
            return []
        
        except requests.exceptions.RequestException as req_err:
            logging.error(f"An error occurred while fetching page {params['page']}: {req_err}")
            return []
        
        except Exception as e:
            logging.error(f"An unexpected error occurred on page {params['page']}: {e}")
            return []

def fetch_orders():
    """Fetch orders from WooCommerce API with improved pagination using concurrency."""
    # Get the current Jalali date
    current_jalali_date = jdatetime.datetime.now()
    from_jalali_date = current_jalali_date - jdatetime.timedelta(days=days_to_fetch)
    jalali_year = from_jalali_date.year
    jalali_month = from_jalali_date.month
    from_jalali_date = jdatetime.datetime(jalali_year, jalali_month, 1)
    print(current_jalali_date, from_jalali_date)
    gregorian_date = from_jalali_date.togregorian()
    gregorian_date
    iso_date_min = gregorian_date.isoformat()

    logging.info(f"Fetching orders since: {iso_date_min}, {convert_to_jalali(iso_date_min)}")

    all_orders = []
    page = 1
    futures = []

    with concurrent.futures.ThreadPoolExecutor() as executor:
        while True:
            # Submit a new task for the current page
            futures.append(executor.submit(fetch_page,
             master_page = "orders",
             params={"after": iso_date_min, "per_page": wc_api['PER_PAGE'], "page": page}))
            page += 1
            
            # Check if we have a limit on the number of pages to fetch
            if page > 10:  # Example: limit to 10 pages
                break

        # Collect results as they complete
        for future in concurrent.futures.as_completed(futures):
            orders = future.result()
            if orders:
                all_orders.extend(orders)

    return all_orders

def create_order_row(order):
    """Create a row of data for an order."""
    billing = order['billing']
    shipping = order['shipping']
    shipping_total = sum(float(item['total']) for item in order['shipping_lines'] if item['total'].replace('.', '', 1).isdigit())

    if order['date_paid']:
        jalali_date, time_part = convert_to_jalali(order['date_paid'])
        date_part = f"{jalali_date.year}/{jalali_date.month:02d}/{jalali_date.day:02d}"
    else:
        date_part = ''
    
    address_1 = shipping['address_1'].translate(persian_to_english)
    address_1 = address_1.translate(arabic_to_english)

    birthday = find_meta_value(order, '_billing_field_529')
    birthday = birthday.translate(persian_to_english)
    birthday = birthday.translate(arabic_to_english)

    return [
        order['id'],
        STATUS.get(order['status']),
        str(date_part),
        order['customer_id'],
        f"{billing['first_name']} {billing['last_name']}",
        str(billing['phone']),
        billing['email'],
        f"{birthday}",
        f"{STATES.get(shipping['state'], shipping['state'])}، {shipping['city']}",
        f"{address_1}",
        str(shipping['postcode']),
        int(order['total']) * 10,
        int(shipping_total) * 10,
        int(order['discount_total']) * 10,
        round(int(order['discount_total']) * 10 / 1.10),
        '', '', '', '', '',
        find_meta_value(order, 'datei'),
        find_meta_value(order, 'marsule'), '', '',
        find_meta_value(order, 'datedeliver')
    ]

def write_products(sheet, line_items):
    """Write product line items to the sheet."""
    for item in line_items:
        product_row = [''] * 15 + [str(item['sku']), item['name'], item['quantity'], int(item['total']) * 10]
        sheet.append(product_row)
        sheet.cell(row=sheet.max_row, column=list(COLUMN_HEADERS.keys()).index("item_total") + 1).number_format = '#,0'
        
        for col in range(list(COLUMN_HEADERS.keys()).index("sepidar_discount") + 1, list(COLUMN_HEADERS.keys()).index("sepidar_id") + 2):
            sheet[sheet.max_row][col - 1].border = borders["right_border"]

def count_integer_rows(sheet, column_name: str, start_row: int, stop_row: int) -> int:
    """
    Count the number of rows in the given sheet that have any integer value in the specified column
    between the start_row and stop_row.

    Args:
        sheet: The Excel sheet to search through.
        column_name: The name of the column to check.
        start_row: The row number to start counting from (1-based index).
        stop_row: The row number to stop counting at (1-based index).

    Returns:
        int: The count of rows that contain any integer in the specified column.
    """
    # Get the index of the column
    headers = [cell.value for cell in sheet[1]]  # Assuming the first row is headers
    if column_name not in headers:
        raise ValueError(f"Column '{column_name}' not found in the sheet.")

    column_index = headers.index(column_name) + 1  # +1 for 1-based index

    count = 0
    # Iterate through rows from start_row to stop_row
    for row in sheet.iter_rows(min_row=start_row, max_row=stop_row):  # Adjust as necessary
        if isinstance(row[column_index - 1].value, int):  # Check if the cell contains an integer
            count += 1

    return count

def add_sum_row(sheet, from_row, last_row, col_list):
    """Add a sum row after the last order of the Jalali month."""
    sum_row_index = last_row + 1
    sheet.insert_rows(sum_row_index)

    for cell in sheet[sum_row_index]:
        cell.border = borders["top_border"]
    
    # Set the label for the sum row
    sheet.cell(row=sum_row_index, column=list(COLUMN_HEADERS.keys()).index('address') + 1, value=TEXT['sum_month_orders_row_text'])  # Adjust column index for label

    lastmonth_orders_count = count_integer_rows(sheet, COLUMN_HEADERS['order_id'], from_row, last_row)
    sheet.cell(row=sum_row_index, column=list(COLUMN_HEADERS.keys()).index('postcode') + 1, value=lastmonth_orders_count)

    """Calculate totals for specified columns in the sheet."""
    column_indices = {name: list(COLUMN_HEADERS.keys()).index(name) + 1 for name in col_list}

    # Add sum formulas for each specified column
    for column_name, column_index in column_indices.items():
        if column_name == 'address' or column_name == 'postcode':
            continue
        sum_formula = f'=SUM({get_column_letter(column_index)}{from_row}:{get_column_letter(column_index)}{last_row})'
        sheet.cell(row=sum_row_index, column=column_index, value=sum_formula)

    # Apply styles to the sum row
    for col in col_list:
        cell = sheet.cell(row=sum_row_index, column=list(COLUMN_HEADERS.keys()).index(col) + 1)
        cell.number_format = '#,0'

    for cell in sheet[sum_row_index]:
        cell.border = borders["all_borders"]
        cell.fill = color["sum_row_bg"]
        if cell.column == 11:
            cell.alignment = Alignment(horizontal='center')
    return sum_row_index

def find_sum_rows(sheet) -> List[List[float]]:
    """Find sum rows in the Excel sheet and return their integer and float values."""
    address_index = list(COLUMN_HEADERS.keys()).index('address')
    sum_row_indices = []
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
        if row[address_index].value == TEXT['sum_month_orders_row_text']: # Adjust this condition based on your identifier
            sum_row_indices.append(row[0].row)  # Append the actual row number
    if not sum_row_indices:
        logging.warning("No sum rows found.")

    return sum_row_indices

def calculate_totals(sheet, sum_rows, col_list):
    # Assuming COLUMN_HEADERS is defined and contains the necessary keys
    column_indices = {name: list(COLUMN_HEADERS.keys()).index(name) + 1 for name in col_list}
    col_total = {}

    # Iterate over each key to construct the formula
    for key in column_indices.keys():
        col_total[key] = '='  # Initialize the formula for each key
        formula_parts = [f"{get_column_letter(column_indices[key])}{i}" for i in sum_rows]  # Create cell references
        col_total[key] += " + ".join(formula_parts)  # Join them into a single formula
    
    return col_total

def append_totals(sheet, col_total, col_list):
    totals_row = ['', '', '', '', '', '', '', '', '', TEXT['sum_all_orders_row_text'], col_total["postcode"],
        col_total["total"], col_total["shipping"], col_total["discount"],
        col_total["sepidar_discount"], '', '', '', col_total["item_total"], '', '', '',
        col_total["com_postal_payment"], col_total["com_postage"]]
    sheet.append(totals_row)

    for cell in sheet[sheet.max_row]:
        cell.border = borders["top_border"]

    for col in col_list:
        cell = sheet.cell(row=sheet.max_row, column=list(COLUMN_HEADERS.keys()).index(col) + 1)
        cell.number_format = '#,0'
    
    for cell in sheet[sheet.max_row]:
        cell.border = borders["all_borders"]
        cell.fill = color["total_sum_bg"]
        if cell.column == 11:
            cell.alignment = Alignment(horizontal='center')

def write_to_excel(excel_file: str, orders: list) -> None:
    """Write JSON data to an Excel file."""
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        logging.warning(f"The file {excel_file} does not exist. A new file will be created.")
        workbook = Workbook()
        sheet = workbook.active
    except IOError as io_err:
        logging.error(f"I/O error occurred while accessing {excel_file}: {io_err}")
        return
    except Exception as e:
        logging.error(f"An unexpected error occurred while opening {excel_file}: {e}")
        return

    if sheet.max_row == 1:
        sheet.append(list(COLUMN_HEADERS.values()))
        logging.info("Write header row.")
        for cell in sheet[sheet.max_row]:
            cell.border = borders["all_borders"]
            cell.fill = color["header_bg"]

    if sheet.max_row > 1:
        sheet.delete_rows(sheet.max_row)
        logging.info(f"Removed last month orders row {sheet.max_row}.")
        sheet.delete_rows(sheet.max_row)
        logging.info(f"Removed total orders row {sheet.max_row}.")

    existing_order_ids = {
        sheet.cell(row=row, column=list(COLUMN_HEADERS.keys()).index("order_id") + 1).value: row
        for row in range(2, sheet.max_row + 1)
    }

    existing_order_ids = {k: v for k, v in existing_order_ids.items() if k is not None}

    last_order_jalali_month = None
    col_list =[
        "address",
        "postcode",
        "total", 
        "shipping", 
        "discount", 
        "sepidar_discount", 
        "item_total", 
        "com_postal_payment", 
        "com_postage"
    ]
    from_row = 2
    sum_row_indices = []
    new_orders_count = 0

    logging.info(f"Total Orders: {len(orders)}, Existed Orders: {len(existing_order_ids)}")

    for order in orders:
        order_id = order['id']
        jalali_date, time_part = convert_to_jalali(order['date_paid'])
        current_order_jalali_month = f"{jalali_date.year}-{jalali_date.month:02d}"
        if last_order_jalali_month == None:
            last_order_jalali_month = current_order_jalali_month

        if order_id in existing_order_ids:
            row_index = existing_order_ids[order_id]

            current_status = order['status']
            lang_existing_status = sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("status") + 1).value
            existing_status = get_key_by_value(STATUS, lang_existing_status)
            
            if existing_status != current_status:
                logging.info(f"Order ID {order_id} status changed from {STATUS.get(existing_status)} to {STATUS.get(current_status)}. Updating...")
                sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("status") + 1).value = STATUS.get(current_status)
                # apply_styles(sheet, row_index, order)  # Reapply styles if necessary
                
            existing_datei = sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("datei") + 1).value
            current_datei = find_meta_value(order, 'datei')

            if current_datei != existing_datei:
                logging.info(f"Order ID {order_id} send date changed from {existing_datei} to {current_datei}. Updating...")
                sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("datei") + 1).value = current_datei
                
            existing_tracking_code = sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("tracking_code") + 1).value
            current_tracking_code = find_meta_value(order, 'marsule')
            if current_tracking_code != existing_tracking_code:
                logging.info(f"Order ID {order_id} tracking code changed from {existing_tracking_code} to {current_tracking_code}. Updating...")
                sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("tracking_code") + 1).value = current_tracking_code

            existing_delivery_date = sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("delivery_date") + 1).value
            current_delivery_date = find_meta_value(order, 'datedeliver')
            if current_delivery_date != existing_delivery_date:
                logging.info(f"Order ID {order_id} delivery date changed from {existing_delivery_date} to {current_delivery_date}. Updating...")
                sheet.cell(row=row_index, column=list(COLUMN_HEADERS.keys()).index("delivery_date") + 1).value = current_delivery_date

            last_order_jalali_month = current_order_jalali_month
            continue

        new_orders_count +=1

        # Check if we need to add a sum row for the previous month
        if current_order_jalali_month != last_order_jalali_month:
            logging.info(f"add a sum row for the {last_order_jalali_month} month")
            sum_row_indices.append(add_sum_row(sheet, from_row, sheet.max_row, col_list))
            from_row = sheet.max_row + 1
            last_order_jalali_month = current_order_jalali_month

        order_row = create_order_row(order)
        sheet.append(order_row)
        row_index = sheet.max_row
        
        logging.info(f"Order ID {order_id} written in Excel row.")
        apply_styles(sheet, row_index, order)
        write_products(sheet, order['line_items'])

        # Calculate com_postage
        com_postage_cell_index = list(COLUMN_HEADERS.keys()).index("com_postage") + 1
        com_postage_value = f"=M{row_index} - W{row_index}"
        sheet.cell(row=row_index, column=com_postage_cell_index, value=com_postage_value)

    # sum_rows = find_sum_rows(sheet)
    # if new_orders_count and last_order_jalali_month and len(sum_rows) > 1 and sum_rows[-1] <  sheet.max_row - 1 :
    #     sheet.delete_rows(sum_rows[-1])
    #     logging.info("Removed last month orders sum row.")

    # Add last month sum row if necessary
    # if len(orders) != len(existing_order_ids):
    # if last_order_jalali_month and new_orders_count > 0:
        # sheet.delete_rows(sum_rows[-1])
    # logging.info("Removed last month orders sum row.")
    sum_rows = find_sum_rows(sheet)
    add_sum_row(sheet, sum_rows[-1] + 1, sheet.max_row, col_list)
    sum_row_indices.append(sheet.max_row)
    logging.info(f"New Orders : {new_orders_count}")
    logging.info("Written last month orders sum row.")
    
    sum_rows = find_sum_rows(sheet)

    col_total = calculate_totals(sheet, sum_rows, col_list)
    append_totals(sheet, col_total, col_list)
    logging.info("Written total orders row.")

    logging.info(f"Saving data to {excel_file}...")

    try:
        workbook.save(excel_file)
        logging.info(f"Data has been appended to {excel_file}.")
    except PermissionError:
        logging.error(f"Permission denied: Unable to save to {excel_file}. Please close the file if it's open.")
    except IOError as io_err:
        logging.error(f"I/O error occurred while saving {excel_file}: {io_err}")
    except Exception as e:
        logging.error(f"Failed to save the Excel file: {e}")

def open_excel_file(file_path):
    """Open the specified Excel file."""
    try:
        if os.name == 'nt':
            os.startfile(file_path)
        elif os.name == 'posix':
            subprocess.call(['xdg-open', file_path])
        else:
            logging.warning("Unsupported OS. Please open the file manually.")
    except FileNotFoundError:
        logging.error(f"File not found: {file_path}. Make sure the file exists.")
    except Exception as e:
        logging.error(f"Error opening the file: {e}")

if __name__ == "__main__":
    # Configure logging with rotation
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            RotatingFileHandler("fetchWCOrders.log", maxBytes=5 * 1024 * 1024, backupCount=2),
            logging.StreamHandler()
        ]
    )

    # Translation tables
    persian_to_english = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
    arabic_to_english = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')

    lang, wc_api, days_to_fetch, color, borders, fonts = load_config('config.json')
    if lang['lang'] == "en":
        from mapping import STATES, ENGLISH_COLUMN_HEADERS, ENGLISH_STATUS, ENGLISH_TEXT
        COLUMN_HEADERS = ENGLISH_COLUMN_HEADERS
        STATUS = ENGLISH_STATUS
        TEXT = ENGLISH_TEXT
    elif lang['lang'] == "fa":
        from mapping import STATES, PERSIAN_COLUMN_HEADERS, PERSIAN_STATUS, PERSIAN_TEXT
        COLUMN_HEADERS = PERSIAN_COLUMN_HEADERS
        STATUS = PERSIAN_STATUS
        TEXT = PERSIAN_TEXT
    else:
        logging.error(f"Missing language configuration in config.json")

    all_orders = fetch_orders()
    # Filter out cancelled orders
    orders = [order for order in all_orders if order['status'] != 'cancelled' and order['status'] != 'pending']

    if orders:
        logging.info("JSON data loaded successfully.")
        orders.sort(key=lambda k: int(k['id']))

        excel_file_path = os.path.join(".", "orders.xlsx")
        write_to_excel(excel_file_path, orders)

        if os.path.exists(excel_file_path):
            open_excel_file(excel_file_path)
        else:
            logging.error("The Excel file was not created.")
    else:
        logging.info("There is no order.")
